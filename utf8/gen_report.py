#!/usr/bin/env python3
"""
gen_report.py  --  Generate an Excel test report from test_suite.sh output.

Usage:
    bash test_suite.sh --no-color 2>&1 | python3 gen_report.py
    bash test_suite.sh --no-color 2>&1 | python3 gen_report.py --out report.xlsx
    python3 gen_report.py --log saved_output.txt --out report.xlsx

One row per TEST GROUP (30 rows matching the 30 test_* functions).
"""

import sys, re, argparse
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl --break-system-packages", file=sys.stderr)
    sys.exit(1)

ap = argparse.ArgumentParser()
ap.add_argument("--log",  default="-",           help="Input (default: stdin)")
ap.add_argument("--out",  default="report.xlsx", help="Output xlsx")
args = ap.parse_args()

raw  = Path(args.log).read_text(errors="replace") if args.log != "-" else sys.stdin.read()
text = re.compile(r'\x1b\[[0-9;]*m').sub("", raw)

META = {
    "File Generation": {
        "category": "File Validation",
        "precond":  "Script executable; test directory writable",
        "steps": (
            "1. Generate a 1000-record structured file.\n"
            "2. Verify exactly 1002 lines (header + 1000 records + footer).\n"
            "3. Verify header string matches expected value.\n"
            "4. Verify footer matches FOOTERTEST########.\n"
            "5. Run -l 500 --dry-run and check show_file_info output contains\n"
            "   File:, Size:, Type: fields and correct line count.\n"
            "6. Create a DOS CRLF file; verify show_file_info reports 'CRLF'."
        ),
        "expected": (
            "1002 lines. Correct header. Footer matches pattern. "
            "show_file_info shows size, line count, file type. CRLF file shows 'CRLF' in Type."
        ),
    },
    "Line Preview (-l)": {
        "category": "Preview",
        "precond":  "1000-line structured file generated",
        "steps": (
            "1. Run: script.sh -l 500 --dry-run --yes\n"
            "2. Verify output contains '500 |' (target line).\n"
            "3. Verify output contains '499 |' (context line before target).\n"
            "4. Verify output contains '501 |' (context line after target)."
        ),
        "expected": "Preview shows lines 499 (context), 500 (target highlighted), 501 (context). File unchanged.",
    },
    "Delete Single Line (-l)": {
        "category": "Delete",
        "precond":  "1000-line structured file generated",
        "steps": (
            "1. Record initial line count (1002).\n"
            "2. Run: script.sh -l 500 --yes --no-color\n"
            "3. Verify line count is now 1001 (reduced by 1).\n"
            "4. Verify footer = FOOTERTEST00000999.\n"
            "5. Verify line containing 'ITEM-0499' absent.\n"
            "6. Verify _modified_* file created."
        ),
        "expected": "Line count -1. Footer decremented to FOOTERTEST00000999. ITEM-0499 absent. _modified_ created.",
    },
    "Header/Footer Protection": {
        "category": "Header/Footer Protection",
        "precond":  "100-line structured file generated",
        "steps": (
            "1. Record initial line count N.\n"
            "2. Run: script.sh -l 1 --yes (attempt header delete).\n"
            "3. Verify line count still N (header silently skipped).\n"
            "4. Run: script.sh -l N --yes (attempt footer delete).\n"
            "5. Verify line count still N (footer silently skipped)."
        ),
        "expected": "Both operations complete. Line count unchanged. Header and footer protected.",
    },
    "Keyword Search (-w)": {
        "category": "Keyword Search",
        "precond":  "1000-line structured file generated",
        "steps": (
            "1. Run: script.sh -w ITEM-0420 --dry-run --yes --no-color\n"
            "2. Verify output contains 'ITEM-0420' (match found in preview).\n"
            "3. Verify output contains 'Matches' (preview section header shown)."
        ),
        "expected": "Match preview shown containing ITEM-0420. File unchanged.",
    },
    "Delete by Keyword (-w)": {
        "category": "Keyword Delete",
        "precond":  "1000-line structured file generated",
        "steps": (
            "1. Record initial line count.\n"
            "2. Run: script.sh -w 'CAT=A' --max-changes 0 --yes\n"
            "3. Verify line count decreased (multiple CAT=A lines removed).\n"
            "4. Verify footer still matches pattern FOOTERTEST########."
        ),
        "expected": "All CAT=A lines removed. Line count reduced. Footer updated and valid.",
    },
    "Position Filter (--pos)": {
        "category": "Position Filter",
        "precond":  "1000-line structured file generated",
        "steps": (
            "1. Run: script.sh -w ITEM --pos 10-20 -n 5 --dry-run --yes\n"
            "2. Verify output contains 'Matches' — search was scoped to\n"
            "   character positions 10-20 and found results."
        ),
        "expected": "Matches found within specified character position range. Search scoped to cols 10-20.",
    },
    "Replace Preview": {
        "category": "Replace",
        "precond":  "11-line test file with ERROR001/ERROR002/ERROR003/ERROR004 lines",
        "steps": (
            "1. Run: script.sh -w ERROR --replace-pos 1-8 --replace-txt RESOLVED --dry-run --yes\n"
            "2. Verify output contains 'Original' (pre-replace lines shown).\n"
            "3. Verify output contains 'Replaced' (post-replace preview shown).\n"
            "4. Verify output contains 'RESOLVED' (replacement text visible)."
        ),
        "expected": "Before/after pairs shown for each ERROR line. File NOT modified (dry-run).",
    },
    "Replace Text Operation": {
        "category": "Replace",
        "precond":  "11-line test file with ERROR001/ERROR002/ERROR003/ERROR004 lines",
        "steps": (
            "1. Record initial line count.\n"
            "2. Run: script.sh -w ERROR --replace-pos 1-8 --replace-txt RESOLVED --yes\n"
            "3. Verify line count unchanged (replace does not remove lines).\n"
            "4. Verify 'RESOLVED' present in file.\n"
            "5. Verify 'ERROR001' absent from file."
        ),
        "expected": "Line count unchanged. Chars 1-8 replaced with RESOLVED. ERROR001 gone.",
    },
    "Replace with Position Search": {
        "category": "Replace",
        "precond":  "11-line test file with ERROR001/ERROR002/ERROR003/ERROR004 lines",
        "steps": (
            "1. Run: script.sh --pos 1-5 -w ERROR --replace-pos 1-8 --replace-txt FIXED___ --yes\n"
            "2. Verify 'FIXED___' present in file.\n"
            "   Note: --pos scopes the search; --replace-pos is the replacement target."
        ),
        "expected": "FIXED___ present. Demonstrates --pos (search scope) and --replace-pos are independent.",
    },
    "Replace Modified Tracking": {
        "category": "Audit Trail",
        "precond":  "11-line test file with ERROR001/ERROR002/ERROR003/ERROR004 lines",
        "steps": (
            "1. Run replace with --yes; verify _modified_* file created.\n"
            "2. Verify _modified_* contains 'ERROR001' (original pre-replace content).\n"
            "3. Re-generate file; run same replace with --no-modified flag.\n"
            "4. Verify no _modified_* file created.\n"
            "5. Verify replacement still applied (RESOLVED present in file)."
        ),
        "expected": "_modified_ contains originals. With --no-modified: no audit file, but replace still applies.",
    },
    "Backup Creation": {
        "category": "Backup",
        "precond":  "100-line structured file; no prior backup",
        "steps": (
            "1. Run: script.sh -l 50 --yes\n"
            "2. Verify _backup file exists.\n"
            "3. Verify _backup has 102 lines (original pre-delete state)."
        ),
        "expected": "_backup created. Backup has 102 lines capturing pre-operation state.",
    },
    "Backup Reuse": {
        "category": "Backup",
        "precond":  "100-line structured file; prior backup may exist",
        "steps": (
            "1. Create backup via first delete; record backup mtime.\n"
            "2. Run a second delete; verify backup mtime unchanged (not overwritten).\n"
            "3. Re-generate file; write a 10-byte truncated fake _backup file.\n"
            "4. Run a delete; verify warning about incomplete/replaced backup.\n"
            "5. Verify backup is now full-size (truncated backup replaced)."
        ),
        "expected": "Normal: backup not overwritten. Truncated backup: warning emitted; replaced with full copy.",
    },
    "Rollback Operation": {
        "category": "Rollback",
        "precond":  "100-line structured file generated",
        "steps": (
            "1. Delete line 50; delete line 60 (two sequential changes).\n"
            "2. Verify modified_lines < original_lines.\n"
            "3. Run: script.sh --rollback --yes\n"
            "4. Verify line count restored to original 102."
        ),
        "expected": "File restored to pre-first-operation state (102 lines). All changes reversed.",
    },
    "Rollback Error Handling": {
        "category": "Rollback",
        "precond":  "100-line structured file; backup file manually deleted",
        "steps": (
            "1. Ensure no _backup file exists.\n"
            "2. Run: script.sh --rollback --yes\n"
            "3. Verify output contains error about backup not found."
        ),
        "expected": "Error message shown. Script exits non-zero.",
    },
    "Regex Matching (--regex)": {
        "category": "Regex",
        "precond":  "1000-line structured file generated",
        "steps": (
            "1. Run: script.sh -w 'ITEM-0[0-9]{3}' --regex --dry-run --yes\n"
            "2. Verify output contains 'ITEM-0' (match found).\n"
            "3. Verify output contains 'Matches' (preview section shown)."
        ),
        "expected": "Regex pattern matches ITEM-0NNN lines. Preview shown. File unchanged.",
    },
    "Regex vs Literal Comparison": {
        "category": "Regex",
        "precond":  "1000-line structured file generated",
        "steps": (
            "1. Run: script.sh -w '[0-9]' --dry-run --yes (no --regex flag).\n"
            "2. Verify output says 'No matches' (literal string not found).\n"
            "3. Run: script.sh -w '[0-9]' --regex --dry-run --yes.\n"
            "4. Verify output shows matches (digit class matches all data lines)."
        ),
        "expected": "Without --regex: no matches. With --regex: many matches. Confirms literal vs regex modes.",
    },
    "Regex Delete": {
        "category": "Regex Delete",
        "precond":  "1000-line structured file generated",
        "steps": (
            "1. Record initial line count.\n"
            "2. Run: script.sh -w 'CAT=A' --regex --max-changes 0 --yes\n"
            "3. Verify line count decreased.\n"
            "4. Verify footer still matches FOOTERTEST########."
        ),
        "expected": "Regex-matched CAT=A lines removed. Footer updated and valid.",
    },
    "Regex Replace": {
        "category": "Regex Replace",
        "precond":  "11-line test file with ERROR001/ERROR002/ERROR003/ERROR004 lines",
        "steps": (
            "1. Run: script.sh -w 'ERROR[0-9]+' --regex --replace-pos 1-8 --replace-txt FIXED___ --yes\n"
            "2. Verify 'FIXED___' present in file.\n"
            "3. Verify 'ERROR001' absent from file."
        ),
        "expected": "Regex matches all ERROR lines. All replaced. ERROR001 gone.",
    },
    "Preview Limit (-n)": {
        "category": "Preview",
        "precond":  "1000-line structured file generated",
        "steps": (
            "1. Run: script.sh -w CAT -n 3 --dry-run --yes --max-changes 0\n"
            "2. Verify output contains '... +N more' overflow indicator\n"
            "   (CAT appears in far more than 3 lines)."
        ),
        "expected": "Overflow indicator '... +N more' shown. Only 3 preview lines displayed.",
    },
    "Dry-Run Mode (--dry-run)": {
        "category": "Dry Run",
        "precond":  "100-line structured file generated",
        "steps": (
            "1. Capture file content before run.\n"
            "2. Run: script.sh -l 50 --dry-run --yes\n"
            "3. Verify file content byte-identical after run.\n"
            "4. Verify line count unchanged and [DRY-RUN] shown in output.\n"
            "5. Run: script.sh -w ITEM --dry-run --yes --max-changes 5\n"
            "   (ITEM matches far more than 5 lines).\n"
            "6. Verify MAX_CHANGES abort message shown even in dry-run.\n"
            "7. Verify file unchanged after aborted dry-run."
        ),
        "expected": "File unchanged. [DRY-RUN] shown. MAX_CHANGES guard aborts on dry-run too (saves scan time).",
    },
    "Modified Line Tracking": {
        "category": "Audit Trail",
        "precond":  "100-line structured file generated",
        "steps": (
            "1. Run: script.sh -l 50 --yes\n"
            "2. Verify _modified_* file exists with exactly 1 line.\n"
            "3. Re-generate; run: script.sh -l 50 --no-modified --yes\n"
            "4. Verify no _modified_* file created.\n"
            "5. Verify line count = 101 (delete still applied)."
        ),
        "expected": "Normal: _modified_ created with 1 line. --no-modified: no audit file, delete still applied.",
    },
    "Edge Cases": {
        "category": "Guards & Edge Cases",
        "precond":  "100-line structured file generated",
        "steps": (
            "1. Run: script.sh -l 99999 --yes → verify 'out of range' error.\n"
            "2. Run: script.sh -f /nonexistent.txt -l 1 → verify 'not found' error.\n"
            "3. Run: script.sh -w ITEM --max-changes 5 --yes → verify abort message\n"
            "   and file line count unchanged.\n"
            "4. Verify script source contains 'validate_allowed_path' and 'realpath'\n"
            "   (path allowlist guard present)."
        ),
        "expected": "Out-of-range: error. No file: error. MAX_CHANGES: abort, file unchanged. Allowlist guard present.",
    },
    "Bulk Operations": {
        "category": "Delete",
        "precond":  "200-line structured file generated",
        "steps": (
            "1. Delete line 100; record count_after1.\n"
            "2. Delete line 99; record count_after2.\n"
            "3. Delete line 98; record count_after3.\n"
            "4. Verify count_after3 < count_after2 < count_after1 < initial."
        ),
        "expected": "Three sequential deletes each reduce line count by 1. Counts strictly decreasing.",
    },
    "Footer Integrity": {
        "category": "Footer Integrity",
        "precond":  "200-line structured file (footer = FOOTERTEST00000200)",
        "steps": (
            "1. Delete line 50.\n"
            "2. Delete line 100.\n"
            "3. Delete line 150.\n"
            "4. Extract numeric part of footer.\n"
            "5. Verify footer numeric value = 197 (200 minus 3 deletions)."
        ),
        "expected": "Footer decrements by 1 per deleted record. After 3 deletions: FOOTERTEST00000197.",
    },
    "Color Output": {
        "category": "Output Formatting",
        "precond":  "100-line structured file generated",
        "steps": (
            "1. Run: script.sh -l 50 --dry-run --yes (no --no-color).\n"
            "2. Verify output contains ANSI escape sequences (\\x1b[...).\n"
            "3. Run: script.sh -l 50 --dry-run --yes --no-color.\n"
            "4. Verify output contains no ANSI escape sequences."
        ),
        "expected": "Default: ANSI codes present. --no-color: no escape codes.",
    },
    "Complex Workflow": {
        "category": "Integration",
        "precond":  "11-line test file with ERROR001/INFO0002/WARN0003 etc lines",
        "steps": (
            "1. Capture original file content.\n"
            "2. Run: script.sh -w INFO --yes → verify INFO absent.\n"
            "3. Run: script.sh -w ERROR --replace-pos 1-8 --replace-txt RESOLVED --yes\n"
            "   → verify RESOLVED present, ERROR prefix absent.\n"
            "4. Verify _backup and _modified_* files exist.\n"
            "5. Run: script.sh --rollback --yes\n"
            "6. Verify file content is byte-identical to original."
        ),
        "expected": "INFO deleted; ERROR replaced; backup and modified files created; rollback restores exact original.",
    },
    "Merge Next Line (-l --merge-next)": {
        "category": "Merge",
        "precond":  "Structured file with one record split across two physical lines",
        "steps": (
            "1. Run --dry-run: verify Merge preview, Line N (kept), Line N+1 (absorbed),\n"
            "   Merged result, and 'footer unchanged' message.\n"
            "2. Verify file unchanged after dry-run.\n"
            "3. Run actual merge: verify line count -1, footer NOT decremented,\n"
            "   merged content correct, absorbed line gone, backup created.\n"
            "4. Attempt merge when next line is footer: verify error.\n"
            "5. Attempt --merge-next with -w instead of -l: verify error.\n"
            "6. DOS CRLF file: run merge; verify no embedded CR in middle of merged line;\n"
            "   verify CR preserved at line endings; verify line count -1."
        ),
        "expected": (
            "Two physical lines joined. Footer unchanged. Surrounding lines intact. "
            "Protected-line and -w misuse rejected. DOS CR preserved after merge."
        ),
    },
    "No Header/Footer Mode (--no-header-footer)": {
        "category": "No-Header Mode",
        "precond":  "Plain 5-line file (apple/banana/cherry/date/elderberry)",
        "steps": (
            "1. Run: script.sh -l 3 --no-header-footer --yes → verify 'cherry' deleted.\n"
            "2. Verify 'banana' and 'elderberry' still present.\n"
            "3. Verify _backup created.\n"
            "4. Run: script.sh -l 1 --no-header-footer --merge-next --dry-run --yes\n"
            "5. Verify Merge preview and Merged result shown (no footer logic triggered)."
        ),
        "expected": "Line deleted without footer update. Merge preview works. No header/footer validation.",
    },
    "UTF-8 Content Handling (Thai focus)": {
        "category": "UTF-8 / Encoding",
        "precond":  "Structured file with Thai (สวัสดี, ราคา), Chinese (中文), Latin-extended (café) content",
        "steps": (
            "1. Literal search -w 'สวัสดี' --dry-run: verify match preview shown.\n"
            "2. Regex search -w '.-ITEM' --regex --dry-run: verify ราคา-ITEM matched.\n"
            "3. Literal search -w '中文' --dry-run: verify Chinese match.\n"
            "4. Literal search -w 'café' --dry-run: verify Latin-extended match.\n"
            "5. Delete สวัสดี line: verify count -1, line absent, ราคา and 中文 intact,\n"
            "   footer decremented.\n"
            "6. Replace ราคา-ITEM line --replace-pos 1-5: verify FIXED present.\n"
            "7. Split-line file: สวัสดี across lines 3-4; --merge-next --dry-run:\n"
            "   verify สวั and สดี fragments shown in preview.\n"
            "8. Run merge: verify สวัสดี recombined, footer unchanged."
        ),
        "expected": (
            "All Thai/Chinese/Latin content searchable, deletable, replaceable. "
            "Merge joins split multibyte record. Footer tracks correctly."
        ),
    },
}

# Parse groups from output
groups_raw = re.split(r'Test Group:\s+', text)[1:]
parsed = []
for block in groups_raw:
    name_end = block.index('\n')
    name     = block[:name_end].strip()
    passes   = re.findall(r'\[PASS\]\s+(.+)', block)
    fails    = re.findall(r'\[FAIL\]\s+(.+?)(?:\s+\(.+?\))?$', block, re.MULTILINE)
    folder_m = re.search(r'Folder:\s+(.+)', block)
    folder   = folder_m.group(1).strip() if folder_m else ""
    timing_m = re.search(r'GROUP\s+\S+\s+(\d+)\s+ms', block)
    timing   = int(timing_m.group(1)) if timing_m else 0
    skipped  = bool(re.search(r'SKIPPED', block))
    parsed.append({"name": name, "folder": folder, "passes": passes,
                   "fails": fails, "timing": timing, "skipped": skipped})

if not parsed:
    print("ERROR: No test groups found in input.", file=sys.stderr)
    sys.exit(2)

# Build workbook
wb = Workbook()
ws = wb.active
ws.title = "Test Report"

HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
PASS_FILL    = PatternFill("solid", fgColor="E2EFDA")
FAIL_FILL    = PatternFill("solid", fgColor="FCE4D6")
SKIP_FILL    = PatternFill("solid", fgColor="FFF2CC")
ALT_FILL     = PatternFill("solid", fgColor="F2F2F2")
PASS_FONT    = Font(name="Arial", bold=True, color="375623", size=10)
FAIL_FONT    = Font(name="Arial", bold=True, color="C00000", size=10)
SKIP_FONT    = Font(name="Arial", bold=True, color="7F6000", size=10)
BODY_FONT    = Font(name="Arial", size=10)
ID_FONT      = Font(name="Arial", bold=True, size=10)
WRAP         = Alignment(wrap_text=True, vertical="top")
CENTER_WRAP  = Alignment(wrap_text=True, vertical="top", horizontal="center")
thin         = Side(style="thin", color="D9D9D9")
BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)

COLS = [
    ("Test ID",          8),
    ("Test Folder",      28),
    ("Category",         18),
    ("Test Description", 32),
    ("Precondition",     28),
    ("Test Steps",       52),
    ("Expected Result",  42),
    ("Automated",        11),
    ("Actual Result",    32),
    ("Status",           10),
    ("Notes",            22),
]

# Title
ws.merge_cells("A1:K1")
c = ws["A1"]
c.value     = f"Reconcile Report Line Editor — Test Report  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}"
c.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
c.fill      = PatternFill("solid", fgColor="1F3864")
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 24

# Header row
for ci, (col_name, col_width) in enumerate(COLS, 1):
    cell = ws.cell(row=2, column=ci, value=col_name)
    cell.font = HEADER_FONT; cell.fill = HEADER_FILL
    cell.alignment = CENTER_WRAP; cell.border = BORDER
    ws.column_dimensions[get_column_letter(ci)].width = col_width
ws.row_dimensions[2].height = 20

# Data rows
for ri, grp in enumerate(parsed, 1):
    er   = ri + 2
    meta = META.get(grp["name"], {})
    status = ("SKIPPED" if grp["skipped"] else
              "FAIL"    if grp["fails"]   else "PASS")
    n_pass = len(grp["passes"])
    n_fail = len(grp["fails"])
    actual = (
        f"{n_pass} assertion(s) passed"
        if status == "PASS" else
        f"{n_fail} FAILED: " + "; ".join(f[:55] for f in grp["fails"][:3])
        if status == "FAIL" else
        "Skipped — requires RUN_LARGE_FILE_TEST=1"
    )
    notes = (f"{grp['timing']} ms" if status != "SKIPPED"
             else "Set RUN_LARGE_FILE_TEST=1 to enable")

    row_data = [
        f"TC-{ri:02d}",
        grp["folder"],
        meta.get("category", "General"),
        grp["name"],
        meta.get("precond",  ""),
        meta.get("steps",    ""),
        meta.get("expected", ""),
        "Yes",
        actual,
        status,
        notes,
    ]

    base_fill = (PASS_FILL if status == "PASS" else
                 FAIL_FILL if status == "FAIL" else SKIP_FILL)

    for ci, value in enumerate(row_data, 1):
        cell = ws.cell(row=er, column=ci, value=value)
        cell.border    = BORDER
        cell.alignment = CENTER_WRAP if ci in (1, 8, 10) else WRAP
        cell.font      = (PASS_FONT if ci == 10 and status == "PASS" else
                          FAIL_FONT if ci == 10 and status == "FAIL" else
                          SKIP_FONT if ci == 10 and status == "SKIPPED" else
                          ID_FONT   if ci == 1 else BODY_FONT)
        cell.fill = base_fill
    ws.row_dimensions[er].height = 90

# Summary
sr = len(parsed) + 4
for label, formula, fnt in [
    ("Summary",  None,                                          Font(name="Arial", bold=True, size=11)),
    ("Total",    f'=COUNTA(A3:A{len(parsed)+2})',              BODY_FONT),
    ("PASS",     f'=COUNTIF(J3:J{len(parsed)+2},"PASS")',     PASS_FONT),
    ("FAIL",     f'=COUNTIF(J3:J{len(parsed)+2},"FAIL")',     FAIL_FONT),
    ("SKIPPED",  f'=COUNTIF(J3:J{len(parsed)+2},"SKIPPED")',  SKIP_FONT),
]:
    row_offset = ["Summary","Total","PASS","FAIL","SKIPPED"].index(label)
    ws.cell(row=sr+row_offset, column=1, value=label).font = fnt
    if formula:
        ws.cell(row=sr+row_offset, column=2, value=formula).font = fnt

ws.freeze_panes = "A3"

out_path = Path(args.out)
wb.save(out_path)

n_pass = sum(1 for g in parsed if not g["skipped"] and not g["fails"])
n_fail = sum(1 for g in parsed if g["fails"])
n_skip = sum(1 for g in parsed if g["skipped"])
print(f"Report written to: {out_path}")
print(f"Groups: {n_pass} PASS  |  {n_fail} FAIL  |  {n_skip} SKIPPED  |  {len(parsed)} total")
sys.exit(0 if n_fail == 0 else 1)